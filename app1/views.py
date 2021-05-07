import os
from twilio.rest import Client


# Your Account Sid and Auth Token from twilio.com/console
# and set the environment variables. See http://twil.io/secure
def twiliocall(request):

    account_sid = 'AC34c392988cee56831dd30b86b6830e7c'
    auth_token = '34b2511286566bc4e444b72a01e06d61'
    client = Client(account_sid, auth_token)

    call = client.calls.create(
                            url='http://demo.twilio.com/docs/voice.xml',
                            to='+918870697143',
                            from_='+18304636211'
                        )

    print(call)
    print('\n')
    print(call.sid)
    return HttpResponse('twilio call')

def twiliomsg(request):

    # account_sid = 'AC34c392988cee56831dd30b86b6830e7c'
    # auth_token = '34b2511286566bc4e444b72a01e06d61'
    # client = Client(account_sid, auth_token)
    #
    # numbers = ['+917603888517', '+918870697143',]
    #
    # for num in numbers:
    #     message = client.messages \
    #         .create(
    #         body='McAvoy or Stewart? These timelines can get so confusing. hiii hii',
    #         from_='+18304636211',
    #         status_callback='http://postb.in/1234abcd',
    #         to = num
    #     )
    #
    # print(message)
    # print('\n')
    # print(message.sid)


    ACCOUNT_SID = 'AC34c392988cee56831dd30b86b6830e7c'
    AUTH_TOKEN = '34b2511286566bc4e444b72a01e06d61'

    numbers = ['+918870697143',]

    # for number in numbers:
    number = ['{"binding_type":"sms", "address":"=%s"}'%i for i in numbers ]



    client = Client(ACCOUNT_SID, AUTH_TOKEN)
    notification = client.notify.services('ISf0fd474511f5e80797823105ff69f8d9') \
        .notifications.create(
        to_binding=number,
        body='Knok-Knok! This is your first Notify SMS')
    print(notification.sid)
    print(notification)


    return HttpResponse('twilio message...')


# def fetch_message():
#     message = client.messages('SM88ce2f6115444b298b7a9201ec55b4ab').fetch()
#
#     print(message.to)
#
#
#
# fetch_message()




















from django.core import signing
from django.shortcuts import render, redirect
from django.http import HttpResponse

from asgiref.sync import sync_to_async
# from channels.db import database_sync_to_async

import asyncio, time, tablib
# import django.async

from .models import *
from .admin import *


def Encrypt(req, pk):
    value = signing.dumps(pk)
    print(value)
    return redirect('encrypt_url', encryptkey=value)


def Decrypt(req, encryptkey):
    print(encryptkey)
    pk = signing.loads(encryptkey)
    data = Customer.objects.get(pk=pk)
    print(data.name)
    return HttpResponse('Data retrieved...')


def Main(req):
    start = time.time()
    # time.sleep(3)
    data1 = Customer.objects.all()
    data2 = Customer.objects.get(pk=128)
    print(data1)
    print(data2)
    total = (time.time() - start)
    print(total)
    return HttpResponse('sync')

@sync_to_async
def AsyncMain(req):
    start_time = time.time()
    # await asyncio.sleep(3)
    # data1 = await sync_to_async(list)(Customer.objects.all())
    # data2 = await sync_to_async(Customer.objects.get)(pk=128)
    data1 = Customer.objects.all()
    data2 = Customer.objects.get(pk=128)
    print(data1)
    print(data2)

    total_time = (time.time() - start_time)
    print(total_time)
    return HttpResponse('async')


def Excel(req):
    if req.method == 'POST':
        # customer_resource = CustomerResource()
        dataset = tablib.Dataset()
        imported_data = dataset.load(req.FILES['excelfile'].read(), format='xlsx')
        for data in imported_data:
            Customer.objects.create(name=data[0], phone=data[1], product_id=data[2])
        # dataset.xlsx = request.FILES['excelfile'].read()
        # customer_resource.import_data(dataset, dry_run=False)
    return render(req, 'index.html')


def FileView(req):
    obj = FileUpload.objects.get(pk=1).audio
    print(obj.path)
    #------for download file--------
    # response = HttpResponse(obj, content_type='audio/mpeg')
    # response['Content-Disposition'] = "attachment; filename=music.mp3"
    # return response
    return render(req, 'redirect.html', {'path':obj.path})


from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from django.http import HttpResponse


def excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "MyTestSheet"

    row_num = 1

    columns = ['Date', 'Branch', 'Product / Unit', 'Openning Stock', ]

    col_num = 0
    for value in columns:
        col_num = col_num + 1
        cell = ws.cell(row=row_num, column=col_num, value=value)
        cell.font = Font(name='Chandas', bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center')
        column_letter = get_column_letter(col_num)
        column_dimensions = ws.column_dimensions[column_letter]
        column_dimensions.width = 20
        ws.row_dimensions[row_num].height = 25
        cell.fill = PatternFill(start_color='00838F',
                                end_color='00838F',
                                fill_type='solid')

    row = [Date.strftime("%d-%m-%Y"),
           Product.objects.get(pk=query['product']).name + " - " + Unit.objects.get(pk=query['unit']).name,
           query['opening_stock']]

    for value in row:
        col_num = col_num + 1
        value_cell = ws.cell(row=row_num, column=col_num, value=value)
        value_cell.alignment = Alignment(horizontal='center')
        value_cell.fill = PatternFill(start_color='E3F2FD',
                                      end_color='E3F2FD',
                                      fill_type='solid')
        if value == 0:
            value_cell.font = Font(color='FF5252')

    row_num += 1
    ws['A' + str(row_num)].value = ' '
    row_num += 1
    ws.cell(row=row_num, column=1, value=" ")
    row_num += 1

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', )
    response['Content-Disposition'] = 'attachment; filename=product_list.xlsx'
    file = wb.save(response)
    return response



